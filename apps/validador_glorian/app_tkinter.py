import os
import re
import glob
import pandas as pd
import subprocess
import calendar
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import traceback
from PIL import Image, ImageDraw, ImageTk
from tkinter import messagebox
from decimal import Decimal

import tkinter as tk
import tkcalendar
import babel.numbers
from tkinter import ttk
from tkinter import filedialog

import customtkinter as ctk
import io
import msoffcrypto
import pdb
import unicodedata

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__),'.env'))

__PASSWORD_CARTEIRA_OFICIAL = os.getenv('PASSWORD_CARTEIRA_OFICIAL')


path_carteira_oficial = ''
path_carteira_glorian = ''


path_carteira_glorian_3 = ''
path_carteira_glorian_4 = ''

path_carteira_glorian_aux = ''
data_ref = None


df_carteira_oficial = pd.DataFrame()
df_carteira_glorian = pd.DataFrame()

selected_tab = None

relacao = {}
# key -> carteira oficial
# valor -> carteira glorian
relacao['Mês'] = 'inicio_fornecimento'
relacao['Data do Fechamento'] = 'Data Acordo Comercial'
relacao['Código'] = 'Código'
relacao['Submercado'] = 'Submercado'
relacao['CNPJ'] = 'CNPJ'
relacao['Operação'] = 'Operação'
relacao['Perfil'] = 'Produto'
# relacao['Unidade de Volume'] = 'Unidade'
# relacao['MWh'] = 'Quantidade (Convertida)'
relacao['Tipo'] = 'TipoPreco'
relacao['MWh'] = 'Qtd vigente (MWh)'
relacao['R$/MWh'] = 'Preço Original'
relacao['DU'] = 'Condição de Pagamento'
relacao['Data Base'] = 'Data Base'
relacao['Data de Reajuste'] = 'Reajuste'
relacao['Índice de Reajuste'] = 'Índice (Reajuste)'
relacao['Flex Inferior'] = '% Flexibilidade Mensal (Mín)'
relacao['Flex Superior'] = '% Flexibilidade Mensal (Máx)'
relacao['Sazo Min'] = '% Sazonalização (Mín)'
relacao['Sazo Max'] = '% Sazonalização (Max)'
# relacao['Modulação'] = 'Modulação de Carga'
relacao['Meio de Fechamento'] = 'Código Ref.(Contraparte)'

relacao_fechamento = {}
relacao_fechamento['BALCAO'] = np.NaN


relacao_submercado = {}
relacao_submercado['SUDESTE'] = 'SE/CO' 
relacao_submercado['SUL'] = 'S' 
relacao_submercado['NORDESTE'] = 'NE'
relacao_submercado['NORTE'] = 'N' 


# relacao_perfil = {}
# relacao_perfil['WXE'] = ['RAIZEN POWER']
# # relacao_perfil['WXE INE'] = ['CQ5']
# relacao_perfil['WXE CQ5'] = ['RAIZEN POWER CQ5']
# relacao_perfil['WXE I5'] = ['RAIZEN POWER I5']
# relacao_perfil['WXE I5 2'] = ['RAIZEN POWER I5 2']
# relacao_perfil['WXE I0'] = ['RAIZEN POWER I0']
# relacao_perfil['WXE I1 3'] = ['RAIZEN POWER I1 3']
# relacao_perfil['WXE I1'] = ['RAIZEN POWER I1']
# relacao_perfil['WXE I1 2'] = ['RAIZEN POWER I1 2']


relacao_perfil = {}
relacao_perfil['RAIZEN POWER'] = ['CONV']

relacao_perfil['RAIZEN POWER CQ5'] = ['CQ5']
relacao_perfil['RAIZEN POWER INE'] = ['CQ5']

relacao_perfil['RAIZEN POWER I5'] = ['INC_50']
relacao_perfil['RAIZEN POWER I5 2'] = ['INC_50_2']
relacao_perfil['RAIZEN POWER I0'] = ['INC_0']
relacao_perfil['RAIZEN POWER I1 3'] = ['INC_100_3']
relacao_perfil['RAIZEN POWER I1'] = ['INC_100']
relacao_perfil['RAIZEN POWER I1 2'] = ['INC_100_2']

relacao_modulacao = {}
relacao_modulacao['CARGA'] = ['C - Perfil de Carga']
relacao_modulacao['DECLARADA'] = ['D - Modulação Declarada']
relacao_modulacao['FLAT'] = ['F - Flat']
relacao_modulacao['CONFORME GERAÇÃO'] = ['G - Geração']


relacao_codigo = {}
relacao_codigo['20817'] = ['127']
relacao_codigo['20818'] = ['291']
relacao_codigo['20804'] = ['277']
relacao_codigo['20885'] = ['299']
relacao_codigo['20819'] = ['110']
relacao_codigo['20820'] = ['128']
relacao_codigo['20834'] = ['129']
relacao_codigo['20807'] = ['132']
relacao_codigo['20805'] = ['135']
relacao_codigo['20816'] = ['137']
relacao_codigo['20808'] = ['143']
relacao_codigo['20872'] = ['182']
relacao_codigo['20877'] = ['256']
relacao_codigo['20879'] = ['278']
relacao_codigo['20878'] = ['280']
# relacao_codigo['20892'] = ['304', '305']
relacao_codigo['20894'] = ['306']
relacao_codigo['20814'] = ['109']
relacao_codigo['20884'] = ['293']
# relacao_codigo['20888'] = ['301', '302']
relacao_codigo['20823'] = ['136']
relacao_codigo['20825'] = ['139']
relacao_codigo['20826'] = ['150']
relacao_codigo['20898'] = ['312']
relacao_codigo['20809'] = ['126']
relacao_codigo['20810'] = ['131']
relacao_codigo['20811'] = ['133']
relacao_codigo['20831'] = ['142']
relacao_codigo['20832'] = ['146']
relacao_codigo['20833'] = ['148']
relacao_codigo['20824'] = ['149']
relacao_codigo['20852'] = ['168']
relacao_codigo['20853'] = ['169']
relacao_codigo['20887'] = ['303']
relacao_codigo['20897'] = ['311']
# relacao_codigo['20895'] = ['313', '314']
relacao_codigo['20827'] = ['141']
relacao_codigo['20829'] = ['145']
relacao_codigo['20828'] = ['144']
relacao_codigo['20830'] = ['147']
relacao_codigo['20644'] = ['38']
relacao_codigo['20645'] = ['39']
relacao_codigo['20813'] = ['113']
relacao_codigo['20812'] = ['106']
relacao_codigo['20642'] = ['24']
relacao_codigo['20643'] = ['37']
relacao_codigo['20822'] = ['111']
relacao_codigo['20821'] = ['112']
relacao_codigo['20835'] = ['152']
relacao_codigo['20836'] = ['153']
relacao_codigo['20837'] = ['154']
relacao_codigo['20839'] = ['155']
relacao_codigo['20838'] = ['156']
relacao_codigo['20843'] = ['158']
relacao_codigo['20842'] = ['159']
relacao_codigo['20844'] = ['160']
relacao_codigo['20846'] = ['161']
relacao_codigo['20845'] = ['162']
relacao_codigo['20848'] = ['163']
relacao_codigo['20847'] = ['164']
relacao_codigo['20841'] = ['165']
relacao_codigo['20840'] = ['166']
relacao_codigo['20851'] = ['167']
relacao_codigo['20850'] = ['170']
relacao_codigo['20854'] = ['171']
relacao_codigo['20855'] = ['172']
relacao_codigo['20857'] = ['173']
relacao_codigo['20856'] = ['174']
relacao_codigo['20849'] = ['175']
relacao_codigo['20860'] = ['176']
relacao_codigo['20859'] = ['177']
relacao_codigo['20858'] = ['178']
relacao_codigo['20861'] = ['179']
relacao_codigo['20862'] = ['180']
relacao_codigo['20871'] = ['181']
relacao_codigo['20868'] = ['185']
relacao_codigo['20869'] = ['186']
relacao_codigo['20870'] = ['187']
relacao_codigo['20865'] = ['190']
relacao_codigo['20866'] = ['230']
# relacao_codigo['20867'] = ['231', '289', '290']
# relacao_codigo['20864'] = ['246', '288']
relacao_codigo['20863'] = ['247']
# relacao_codigo['20875'] = ['254', '255']
relacao_codigo['20883'] = ['292']
relacao_codigo['20880'] = ['295']
relacao_codigo['20881'] = ['296']
relacao_codigo['20882'] = ['297']
relacao_codigo['20710'] = ['56']
relacao_codigo['20788'] = ['89']
relacao_codigo['20789'] = ['90']
relacao_codigo['20547'] = ['15']
relacao_codigo['20602'] = ['17']
relacao_codigo['20597'] = ['36']
relacao_codigo['20662'] = ['40']
relacao_codigo['20806'] = ['134']
relacao_codigo['20896'] = ['314']
relacao_codigo['20900'] = ['331']
relacao_codigo['20886'] = ['298']




colunas_planilha_saida = [f"{ref} | {relacao[ref]}" for ref in relacao]
df_diferencas = ""
df_mensagens = ""



def verificar_arquivos_entrada_aba_1e3():
    global path_carteira_oficial
    global path_carteira_glorian

    path_home =  os.path.expanduser("~")
    
    patterns_path_carteira = os.path.join(path_home, 'Desktop', 'carteira wxe*')
    paths_carteira_oficial = glob.glob(patterns_path_carteira)
    if len(paths_carteira_oficial) > 0:
        paths_carteira_oficial.sort(reverse=True)
        path_carteira_oficial = paths_carteira_oficial[0]
    else:
        patterns_path_carteira = os.path.join(path_home, 'Downloads', 'carteira wxe*')
        paths_carteira_oficial = glob.glob(patterns_path_carteira)
        if len(paths_carteira_oficial) > 0:
            paths_carteira_oficial.sort(reverse=True)
            path_carteira_oficial = paths_carteira_oficial[0]
    

    if os.path.exists(path_carteira_oficial):
        entrada_caminho1.delete(0, tk.END)  # Limpa o campo de entrada
        entrada_caminho1.insert(0, path_carteira_oficial)  # Insere o caminho do arquivo selecionado
        entrada_caminho5.delete(0, tk.END)  # Limpa o campo de entrada
        entrada_caminho5.insert(0, path_carteira_oficial)  # Insere o caminho do arquivo selecionado

    path_carteira_glorian_local = os.path.join(path_home, 'Downloads', 'Portifólio_(V2)_*.csv')
    arquivos = glob.glob(path_carteira_glorian_local)
    arquivos.sort(reverse=True)
    if len(arquivos) > 0:
        path_carteira_glorian = os.path.join(arquivos[0])
        entrada_caminho2.delete(0, tk.END)  # Limpa o campo de entrada
        entrada_caminho2.insert(0, path_carteira_glorian)  # Insere o caminho do arquivo selecionado
        entrada_caminho6.delete(0, tk.END)  # Limpa o campo de entrada
        entrada_caminho6.insert(0, path_carteira_glorian)  # Insere o caminho do arquivo selecionado
        
        
def verificar_arquivos_entrada_aba2():
    
    path_home =  os.path.expanduser("~")
    path_carteira_glorian_local = os.path.join(path_home, 'Downloads', 'Portifólio_(V2)_*.csv')
    arquivos = glob.glob(path_carteira_glorian_local)
    arquivos.sort(reverse=True)
    if len(arquivos) > 0:
        path_carteira_glorian_3 = os.path.join(arquivos[0])
        path_carteira_glorian_4 = os.path.join(arquivos[1])
        entrada_caminho3.delete(0, tk.END)  # Limpa o campo de entrada
        entrada_caminho3.insert(0, path_carteira_glorian_3)  # Insere o caminho do arquivo selecionado
        entrada_caminho4.delete(0, tk.END)  # Limpa o campo de entrada
        entrada_caminho4.insert(0, path_carteira_glorian_4)  # Insere o caminho do arquivo selecionado
        


def mostrar_dados_carteira_glorian():

    try:
        resetando_statusAplicacao()
        rotulo_status.configure(text="Comparando  ", image=icone_amarelo, compound="right")
        janela.update_idletasks()
        
        global df_carteira_oficial
        global df_carteira_glorian
        global df_diferencas
        global df_mensagens
        global __PASSWORD_CARTEIRA_OFICIAL
        global path_carteira_oficial
        global path_carteira_glorian


        df_diferencas = pd.DataFrame(columns = colunas_planilha_saida)
        df_mensagens = pd.DataFrame()

        data_ref1 = date1.get_date()
        data_ref1 = datetime(data_ref1.year, data_ref1.month, data_ref1.day)
        
        data_ref2 = date2.get_date()
        data_ref2 = datetime(data_ref2.year, data_ref2.month, data_ref2.day)

        new_path_carteira_oficial = entrada_caminho1.get()
        new_path_carteira_glorian = entrada_caminho2.get()
        
        if new_path_carteira_glorian != '' and new_path_carteira_oficial != '': 
            
            _, extensao = os.path.splitext(new_path_carteira_oficial)
            if extensao != '.xlsb':
                rotulo_status.configure(text=f"Erro ",image=icone_vermelho, compound="right")
                messagebox.showerror("Erro ", "Ocorreu um erro: Carteira no formato diferente de xlsb")
                returnnew_path_carteira_oficial
                
            path_carteira_oficial = new_path_carteira_oficial
            path_carteira_glorian = new_path_carteira_glorian
            
            if (df_carteira_glorian.empty or path_carteira_glorian):
                rotulo_status.configure(text="Abrindo a carteira glorian", image=icone_amarelo, compound="right")
                janela.update_idletasks()
                df_carteira_glorian = load_carteira_glorian(path_carteira_glorian)

            if (df_carteira_oficial.empty):
                rotulo_status.configure(text="Abrindo a carteira oficial", image=icone_amarelo, compound="right")
                janela.update_idletasks()
                df_carteira_oficial = load_carteira_oficial(path_carteira_oficial)

            compara_carteira_glorian(df_carteira_oficial,df_carteira_glorian, data_ref1, data_ref2)

    except Exception as e:
        logError = traceback.format_exc()
        rotulo_status.configure(text=f"Erro ",image=icone_vermelho, compound="right")
        messagebox.showerror("Erro ", "Ocorreu um erro: " + str(logError))
        
def mostrar_dados_glorian_glorian():
    global path_carteira_glorian_3
    global path_carteira_glorian_4

    try:
        resetando_statusAplicacao()
        rotulo_status.configure(text="Comparando  ", image=icone_amarelo, compound="right")
        janela.update_idletasks()
        global df_diferencas
        global df_mensagens

        df_diferencas = pd.DataFrame(columns = colunas_planilha_saida)
        df_mensagens = pd.DataFrame()

        data_ref1 = date1.get_date()
        data_ref1 = datetime(data_ref1.year, data_ref1.month, data_ref1.day)
        
        data_ref2 = date2.get_date()
        data_ref2 = datetime(data_ref2.year, data_ref2.month, data_ref2.day)

        
        path_carteira_glorian_3 = entrada_caminho3.get()
        path_carteira_glorian_4 = entrada_caminho4.get()

        if path_carteira_glorian_3 and path_carteira_glorian_4 :
            
            if os.path.basename(path_carteira_glorian_4) ==  os.path.basename(path_carteira_glorian_3):
                
                janela.update_idletasks()
                rotulo_status.configure(text="Não é valido comparar o mesmo Portifólio!",image=icone_vermelho, compound="right")
                return

            df_glorian_d0 = load_carteira_glorian(path_carteira_glorian_3)
            df_glorian_d1 = load_carteira_glorian(path_carteira_glorian_4)
            compara_glorian_glorian(df_glorian_d0,df_glorian_d1)

    except Exception as e:
        logError = traceback.format_exc()
        rotulo_status.configure(text=f"Erro ",image=icone_vermelho, compound="right")
        messagebox.showerror("Erro ", "Ocorreu um erro: " + str(logError))
        
        
        
        
        
def mostrar_dados_carteira_glorian_pegandoValorPor_Mes_Ano(data_formatada):
    
    
    #dividindo a data pela "/" para pegar mes e ano selecionado
    parts = data_formatada.split('/')
    
    if len(parts) == 2:
        mes, ano = int(parts[0]), int(parts[1])
    else:
        print("data Invalida")

    try:
        resetando_statusAplicacao()
        rotulo_status.configure(text="Comparando  ", image=icone_amarelo, compound="right")
        janela.update_idletasks()
        
        global df_carteira_oficial
        global df_carteira_glorian
        global df_diferencas
        global df_mensagens
        global __PASSWORD_CARTEIRA_OFICIAL

        df_diferencas = pd.DataFrame(columns = colunas_planilha_saida)
        df_mensagens = pd.DataFrame()
        
                
        path_carteira_oficial_5 = entrada_caminho5.get()
        path_carteira_glorian_6 = entrada_caminho6.get()

        if path_carteira_glorian_6 and path_carteira_oficial_5:

            if (df_carteira_glorian.empty or path_carteira_glorian_6):
                df_carteira_glorian = load_carteira_glorian(path_carteira_glorian_6)
                
            if (df_carteira_oficial.empty):
                df_carteira_oficial = load_carteira_oficial(path_carteira_oficial_5)

            compara_carteira_glorian_peloMes_Ano(df_carteira_oficial,df_carteira_glorian, mes, ano)

    except Exception as e:
        logError = traceback.format_exc()
        rotulo_status.configure(text=f"Erro ",image=icone_vermelho, compound="right")
        messagebox.showerror("Erro ", "Ocorreu um erro: " + str(logError))        



def selecionar_arquivo(entrada_caminho,param=''):
    
    if param == 'oficial':
        arquivo = filedialog.askopenfilename(title = "Planilha controle Raízen", initialdir=os.path.expanduser("~/Downloads"), filetypes=[("Excel files", "")])
    else:
        arquivo = filedialog.askopenfilename(title = "Planilha controle Raízen", initialdir=os.path.expanduser("~/Downloads"), filetypes=[("Excel files", "Portifólio_(V2)_*.csv")])
    entrada_caminho.delete(0, tk.END)  # Limpa o campo de entrada
    entrada_caminho.insert(0, arquivo)  # Insere o caminho do arquivo selecionado
    resetando_statusAplicacao()




def load_carteira_oficial(path_carteira_oficial):
    

    from pyxlsb import convert_date

    try:
        # Tentar ler o arquivo Excel sem senha
        carteira_oficial = pd.read_excel(path_carteira_oficial, sheet_name='Carteira', engine='pyxlsb')

    except Exception:    
        # Se ocorrer um erro ao tentar ler sem senha, verificar se a senha é necessária
        decrypted_workbook = io.BytesIO()
        with open(path_carteira_oficial, 'rb') as file:
            office_file = msoffcrypto.OfficeFile(file)
            office_file.load_key(password=__PASSWORD_CARTEIRA_OFICIAL)
            office_file.decrypt(decrypted_workbook)

        # Tentar ler o arquivo Excel com senha
        carteira_oficial = pd.read_excel(decrypted_workbook, sheet_name='Carteira')
    
    colunas_converter_data = ['Mês', 'Data do Fechamento', 'Data de Pagamento', 'Data Base', 'Data de Reajuste']
    for col in colunas_converter_data:
        carteira_oficial[col] = carteira_oficial[col].apply(lambda x: convert_date(x) if pd.notna(x) else np.nan)
        
    return carteira_oficial


def load_carteira_glorian(path_carteira_glorian):

    carteira_glorian = pd.read_csv(path_carteira_glorian, 
                                        encoding="latin", 
                                        delimiter=";",
                                        decimal=',',
                                        thousands='.',
                                        low_memory=False)
    
    if 'Código Ref.(Contraparte)' not in carteira_glorian.columns:
        rotulo_status.configure(text="Carteira do Glorian não apresenta a coluna 'Código Ref.(Contraparte)'",image=icone_vermelho, compound="right")
        raise Exception("Carteira do Glorian não apresenta a coluna 'Código Ref.(Contraparte)'")
    
    return carteira_glorian


def tratar_colunas_oficial(recorte_carteira_oficial):
    
    recorte_carteira_oficial['Data Base'] = \
    recorte_carteira_oficial.apply(lambda x:
                                    x['Data do Fechamento'] 
                                    if str(x['Data Base']) == 'NaT' 
                                    else x['Data Base'], 
                                    axis=1
                                    )

    recorte_carteira_oficial['Data de Reajuste'] = \
    recorte_carteira_oficial['Data de Reajuste'].apply(lambda x: 
                                                    x.strftime("%b/%Y") 
                                                    if str(x) != 'NaT' 
                                                    else np.NaN
                                                    )
    
    

    recorte_carteira_oficial[['Flex Inferior','Flex Superior','Sazo Min','Sazo Max']] =\
    recorte_carteira_oficial[['Flex Inferior','Flex Superior','Sazo Min','Sazo Max']].fillna(0)*100

    
    recorte_carteira_oficial['Submercado'] = \
    recorte_carteira_oficial['Submercado'].str.upper().replace(relacao_submercado)

    recorte_carteira_oficial['Perfil'] =\
    recorte_carteira_oficial['Perfil'].str.upper().map(relacao_perfil).apply(lambda x: ', '.join(x) if x else '')
     
    
    # recorte_carteira_oficial['Meio de Fechamento'] = \
    # recorte_carteira_oficial['Meio de Fechamento'].str.upper().replace(relacao_fechamento) 
    

    # Aplica a remoção de acentos à coluna 'Meio de Fechamento' na mesma linha
    recorte_carteira_oficial['Meio de Fechamento'] = recorte_carteira_oficial['Meio de Fechamento'].fillna('')
    
    recorte_carteira_oficial['Meio de Fechamento'] = \
    recorte_carteira_oficial['Meio de Fechamento'].apply(lambda texto: ''
                                                         .join(char for char in 
                                                               unicodedata.normalize('NFD', texto) 
                                                               if unicodedata.category(char) != 'Mn')
                                                         )
    # Converte a coluna para maiúsculas
    recorte_carteira_oficial['Meio de Fechamento'] = recorte_carteira_oficial['Meio de Fechamento'].str.upper()
    

    recorte_carteira_oficial['Meio de Fechamento'] = \
    recorte_carteira_oficial['Meio de Fechamento'].replace(relacao_fechamento) 
    
    
    recorte_carteira_oficial['Modulação'] = \
    recorte_carteira_oficial['Modulação'].str.upper().replace(relacao_modulacao)\
                                                    .fillna('F - Flat')\
                                                    .apply(lambda x: 
                                                    'C - Perfil de Carga' 
                                                    if x == 'CONFORME CARGA'
                                                    else x)

    recorte_carteira_oficial['Operação'] = \
    recorte_carteira_oficial['Operação'].str.capitalize()

    recorte_carteira_oficial['R$/MWh'] = \
    recorte_carteira_oficial['R$/MWh'].round(2)


    try:
        recorte_carteira_oficial['Valor Nota sem ICMS']=\
        recorte_carteira_oficial['Valor Nota sem ICMS'].abs().astype(float).round(2)
    except:
        pdb.set_trace()

    prefixoa_codigos_desiconsiderados = ['Ajuste','Carga','COGEN','GERENCIAL','Gerencial']
    for prefixo in prefixoa_codigos_desiconsiderados:
        filtro = recorte_carteira_oficial['Código'].astype(str).str.startswith(prefixo)
        recorte_carteira_oficial = recorte_carteira_oficial[~filtro].copy()
    
    recorte_carteira_oficial['Código'] = recorte_carteira_oficial['Código'].astype(str)

    # recorte_carteira_oficial['Código'] = pd.to_numeric(recorte_carteira_oficial['Código'], errors='coerce').astype('Int64')

    # recorte_carteira_oficial['Código'] = recorte_carteira_oficial['Código'].fillna('')
    # recorte_carteira_oficial['Código'] = recorte_carteira_oficial['Código'].astype(str)
    # recorte_carteira_oficial['Código'] = recorte_carteira_oficial['Código'].astype(str).str.replace(",","", regex=False).str.replace(".","", regex=False)
    # recorte_carteira_oficial['Código'] = recorte_carteira_oficial['Código'].replace(relacao_codigo).str.upper()

    # recorte_carteira_oficial = recorte_carteira_oficial[~recorte_carteira_oficial['Código'].str.contains(r'Cargas\d+\sOUT\d+', case=False, regex=True)]
    # recorte_carteira_oficial['Código'] = pd.to_numeric(recorte_carteira_oficial['Código'], errors='coerce').astype('Int64')
    
    # recorte_carteira_oficial['Código'] = recorte_carteira_oficial['Código'].astype(str).replace('', '0')

    
    # recorte_carteira_oficial['Código'] = recorte_carteira_oficial['Código'].apply(lambda x: re.search(r'\d{5}', str(x)).group() if re.search(r'\d{5}', str(x)) else '')
    # recorte_carteira_oficial['Código'].fillna(9999999,inplace=True)
    # recorte_carteira_oficial['Código'] = recorte_carteira_oficial['Código'].apply(Decimal)

    return recorte_carteira_oficial


def tratar_colunas_glorian(recorte_carteira_glorian):
    
    global selected_tab


       
    recorte_carteira_glorian['inicio_fornecimento'] = pd.to_datetime(
                                                                    recorte_carteira_glorian['Ano'].astype(str) + 
                                                                    '/' + 
                                                                    recorte_carteira_glorian['Mês'].astype(str) + 
                                                                    '/01'
                                                                    )

    if selected_tab == 0:
        recorte_carteira_glorian['Data Acordo Comercial'] = \
        recorte_carteira_glorian['Data Acordo Comercial'].apply(lambda x: \
                                                            x if str(x) != 'nan' \
                                                            else 'nan'
                                                            )
    else:
        recorte_carteira_glorian['Data Acordo Comercial'] = \
        recorte_carteira_glorian['Data Acordo Comercial'].apply(lambda x: 
                                                        datetime\
                                                        .strptime(x, '%d/%m/%Y')
                                                        if (str(x) != 'nan') 
                                                        else 'nan'
                                                        )
    
    
    recorte_carteira_glorian['Início Fornecimento'] = \
    recorte_carteira_glorian['Início Fornecimento'].apply(lambda x: 
                                                    datetime\
                                                    .strptime(x, '%d/%m/%Y')
                                                    if (str(x) != 'nan') 
                                                    else 'nan'
                                                    )
    
    recorte_carteira_glorian['Final Fornecimento'] = \
    recorte_carteira_glorian['Final Fornecimento'].apply(lambda x: 
                                                    datetime\
                                                    .strptime(x, '%d/%m/%Y')
                                                    if (str(x) != 'nan') 
                                                    else 'nan'
                                                    )
    

    recorte_carteira_glorian['Data Base'] = \
    recorte_carteira_glorian['Data Base'].apply(lambda x: 
                                            datetime.strptime(x, '%d/%m/%Y')\
                                            .strftime("%b/%Y")
                                            if (str(x) != 'nan') 
                                            else 'nan'
                                            )

    recorte_carteira_glorian['Reajuste'] = \
    recorte_carteira_glorian['Reajuste'].apply(lambda x:
                                            datetime\
                                            .strptime(x, '%d/%m/%Y')\
                                            .strftime("%b/%Y") 
                                            if (str(x) != 'nan') 
                                            else x 
                                            )

    recorte_carteira_glorian['Código'] = \
    recorte_carteira_glorian['Código'].str.replace(
                                                'RAIZEN-|WX-|-C|-V|CSG-|WBC-|CUST-RAIZEN|_.*|', "",
                                                regex=True)
    
    recorte_carteira_glorian['Código'] = \
    recorte_carteira_glorian['Código'].replace('', '0', regex=True)  # Substitui strings vazias por '0'
    
    # recorte_carteira_glorian['Código'] = \
    # recorte_carteira_glorian['Código'].str.replace(r'[^0-9]', '', regex=True).apply(Decimal)
    
    
    filtro_pld_ou_preco_fixo = recorte_carteira_glorian['Preço Original'].str.lower().str.contains('pld')
    recorte_carteira_glorian['TipoPreco'] = np.nan
    recorte_carteira_glorian.loc[filtro_pld_ou_preco_fixo,'TipoPreco'] = 'pld+'
    recorte_carteira_glorian.loc[~filtro_pld_ou_preco_fixo,'TipoPreco'] = 'fixo'
    
    
    recorte_carteira_glorian['Preço Original'] =\
    pd.to_numeric(recorte_carteira_glorian['Preço Original']\
                                        .replace('', 'nan')\
                                        .str.replace('PLD {0,}[+, ]', '', regex=True)\
                                        .str.replace(',', '.', regex=True)\
                                        .str.replace(' ', '', regex=True),
                                        errors='coerce'
                                        )

    recorte_carteira_glorian[['% Sazonalização (Mín)','% Flexibilidade Mensal (Mín)']]=\
    recorte_carteira_glorian[['% Sazonalização (Mín)','% Flexibilidade Mensal (Mín)']]*-1
    
    return recorte_carteira_glorian


    
def diferenca_Net_glorian(planilha_d0,planilha_d1,grupo,str_date_d0,str_date_d1):
    
        if grupo == 'mes':
            grupo = 'inicio_fornecimento'
        else:
            grupo ='Ano'

        
        net_total_d0 = planilha_d0.groupby([grupo])['Qtd vigente (MWh)'].sum()
        net_total_d1 = planilha_d1.groupby([grupo])['Qtd vigente (MWh)'].sum()

        dif_net_total = pd.DataFrame(net_total_d0 - net_total_d1).round()
        
        datas_com_dif = dif_net_total[dif_net_total['Qtd vigente (MWh)']!=0].index

        net_total_d1 = pd.DataFrame(net_total_d1)
        net_dif_total_d1=net_total_d1[net_total_d1.index.isin(datas_com_dif)]
        
        net_total_d0 = pd.DataFrame(net_total_d0)
        net_dif_total_d0 = net_total_d0[net_total_d0.index.isin(datas_com_dif)]
        
        datas_com_dif = net_dif_total_d0.index.intersection(net_dif_total_d1.index)
        if datas_com_dif.empty:
            janela.update_idletasks()
            rotulo_status.configure(text="Não tem diferenças Qtd vigente (MWh) para essas datas!!  ",image=icone_vermelho, compound="right")
            return
        if grupo =="Ano":
            
            net_ano_d1 = pd.DataFrame(net_dif_total_d1.groupby(['Ano'])['Qtd vigente (MWh)'].sum())
            net_ano_d1=net_ano_d1.reset_index()
            net_ano_d1[f'Net Ano (MWm) {str_date_d1}']=net_ano_d1.apply(lambda x: x['Qtd vigente (MWh)']/8784 if calendar.isleap(x['Ano']) 
                                                        else x['Qtd vigente (MWh)']/8760, axis=1)
            net_ano_d1 = net_ano_d1.set_index('Ano')\
                .drop('Qtd vigente (MWh)',axis=1)
            
            net_ano_d0 = pd.DataFrame(net_dif_total_d0.groupby(['Ano'])['Qtd vigente (MWh)'].sum())
            net_ano_d0=net_ano_d0.reset_index()
            net_ano_d0[f'Net Ano (MWm) {str_date_d0}']=net_ano_d0.apply(lambda x: x['Qtd vigente (MWh)']/8784 if calendar.isleap(x['Ano']) 
                                                        else x['Qtd vigente (MWh)']/8760, axis=1)
            net_ano_d0 = net_ano_d0.set_index('Ano')\
                .drop('Qtd vigente (MWh)',axis=1)

            dif_net_d0_d1 = pd.concat([net_ano_d0,net_ano_d1],axis=1)
            dif_net_d0_d1[f"DIF NET (MWm)"] = net_ano_d0[f'Net Ano (MWm) {str_date_d0}'] - net_ano_d1[f'Net Ano (MWm) {str_date_d1}']
        
        if grupo == 'inicio_fornecimento':
            net_medio_d0 = planilha_d0.groupby(['inicio_fornecimento'])['Quantidade (Net)'].sum()
            net_medio_d0_meses = net_medio_d0[net_medio_d0.index.isin(datas_com_dif)==True]
            
            net_medio_d1 = planilha_d1.groupby(['inicio_fornecimento'])['Quantidade (Net)'].sum()
            net_medio_d1_meses = net_medio_d1[net_medio_d1.index.isin(datas_com_dif)==True]
            dif_net_medio = pd.DataFrame(net_medio_d0_meses - net_medio_d1_meses)
            dif_net_medio=dif_net_medio.rename(columns= {'Quantidade (Net)':'DIF NET (MWm)'})

            net_dif_total_d0 = net_dif_total_d0.rename(columns= {'Qtd vigente (MWh)':f'Qtd vigente(MWh) ({str_date_d0})'})
            net_dif_total_d1 = net_dif_total_d1.rename(columns= {'Qtd vigente (MWh)':f'Qtd vigente(MWh) ({str_date_d1})'})

            dif_net_d0_d1 = pd.concat([net_dif_total_d0,net_dif_total_d1],axis=1)
            dif_net_d0_d1[f"DIF NET (MWh)"] = dif_net_d0_d1[f'Qtd vigente(MWh) ({str_date_d0})'] - dif_net_d0_d1[f'Qtd vigente(MWh) ({str_date_d1})']
            dif_net_d0_d1 = pd.concat([dif_net_d0_d1,dif_net_medio],axis=1)

            dif_net_d0_d1 = dif_net_d0_d1.sort_index()
            dif_net_d0_d1.index = dif_net_d0_d1.index.strftime("%b/%y")
        return dif_net_d0_d1


def diferenca_operacoes_glorian(planilha_d0,planilha_d1,tipo,str_date_d0,str_date_d1):
    
        if tipo == 'novas': 
            
            planilha_d0_filtrada = planilha_d0[planilha_d0['Data Acordo Comercial'] >= datetime.strptime(str_date_d1,"%d-%b-%y")] 
            planilha_d1_filtrada = planilha_d1[planilha_d1['Data Acordo Comercial'] >= datetime.strptime(str_date_d1,"%d-%b-%y")]
            
            valores_contidos = np.isin(planilha_d0_filtrada.values, planilha_d1_filtrada.values)
            planilha_d0_filtrada_novas = planilha_d0_filtrada[~valores_contidos].drop_duplicates()

            operacoes = planilha_d0_filtrada_novas.groupby(['Código','Contraparte','CNPJ','Operação','Produto','Submercado'])['Qtd vigente (MWh)'].sum()
            operacoes = operacoes.to_frame()
            operacoesFornecimento = planilha_d0_filtrada_novas.groupby(['Código']).agg({'Início Fornecimento':np.min,'Final Fornecimento':np.max})

            operacoes = operacoes.join(operacoesFornecimento[['Início Fornecimento','Final Fornecimento']])
            operacoes['Início Fornecimento'] = operacoes['Início Fornecimento'].dt.strftime("%b/%Y")
            operacoes['Final Fornecimento'] = operacoes['Final Fornecimento'].dt.strftime("%b/%Y")
            
                    
        else:
            planilha_d0_filtrada = planilha_d0[planilha_d0['Data Acordo Comercial'] < datetime.strptime(str_date_d0,"%d-%b-%y")]
            planilha_d1_filtrada = planilha_d1[planilha_d1['Data Acordo Comercial'] < datetime.strptime(str_date_d0,"%d-%b-%y")]
            
            soma_detalhada_d0 = planilha_d0_filtrada.groupby(['inicio_fornecimento','Código','Contraparte','CNPJ','Operação','Produto','Submercado'])['Qtd vigente (MWh)'].sum()
            soma_detalhada_d1 = planilha_d1_filtrada.groupby(['inicio_fornecimento','Código','Contraparte','CNPJ','Operação','Produto','Submercado'])['Qtd vigente (MWh)'].sum()
            
            
            diferenca_detalhada = pd.DataFrame(soma_detalhada_d0 - soma_detalhada_d1).round()
            
            diferenca_detalhada = diferenca_detalhada[(diferenca_detalhada['Qtd vigente (MWh)'] != 0) & (diferenca_detalhada['Qtd vigente (MWh)'].isna() == False)]
            diferenca_detalhada_index = diferenca_detalhada.index

            soma_detalhada_com_diferenca_d0 = pd.DataFrame(soma_detalhada_d0[soma_detalhada_d0.index.isin(diferenca_detalhada_index)])
            soma_detalhada_com_diferenca_d0=soma_detalhada_com_diferenca_d0.rename(columns= {'Qtd vigente (MWh)':f'Qtd vigente(MWh) ({str_date_d0})'})

            soma_detalhada_com_diferenca_d1 = pd.DataFrame(soma_detalhada_d1[soma_detalhada_d1.index.isin(diferenca_detalhada_index)])
            soma_detalhada_com_diferenca_d1=soma_detalhada_com_diferenca_d1.rename(columns= {'Qtd vigente (MWh)':f'Qtd vigente(MWh) ({str_date_d1})'})

            operacoes = pd.concat([soma_detalhada_com_diferenca_d0,soma_detalhada_com_diferenca_d1,diferenca_detalhada],axis=1)
            operacoes.index = operacoes.index.set_levels(operacoes.index.levels[0].strftime('%b/%y'), level=0)
        
        if not operacoes.empty:
            operacoes = operacoes.sort_index(level=0)
        
        return operacoes

            
def style_sheets(path_dile_saida):
    
        workbook = load_workbook(filename=path_dile_saida)
        for worksheet in workbook:
            for column in worksheet.columns:
                max_length = 0
                column_name = column[0].column_letter
                for cell in column:

                    try:
                        if isinstance(cell.value, (int,float)) and column_name not in ['A','B'] and worksheet != workbook['NET com diferença']:
                            cell.number_format = '#,##0.000'    
                        
                        if isinstance(cell.value, (int,float)) and worksheet == workbook['NET com diferença'] and column_name not in ['H']:
                            cell.number_format = '#,##0.000' 

                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_name].width = adjusted_width
        workbook.save(filename=path_dile_saida)

    
def compara_carteira_glorian(df_carteira_oficial,df_carteira_glorian, data_inicial, data_final):
    
        janela.update_idletasks()
        
        global df_diferencas
        global df_mensagens

        filtro1 = df_carteira_oficial['Data do Fechamento']>=data_inicial
        filtro2 = df_carteira_oficial['Data do Fechamento']<=data_final
        
        
        df_carteira_glorian['Data Acordo Comercial'] = pd.to_datetime(df_carteira_glorian['Data Acordo Comercial'], format='%d/%m/%Y')
        
        filtro3 = df_carteira_glorian['Data Acordo Comercial']>=data_inicial
        filtro4 = df_carteira_glorian['Data Acordo Comercial']<=data_final
            
        recorte_carteira_oficial = df_carteira_oficial[filtro1 & filtro2].copy()
        recorte_carteira_glorian = df_carteira_glorian[filtro3 & filtro4].copy()

        if not recorte_carteira_glorian.empty and not recorte_carteira_oficial.empty:

            recorte_carteira_oficial= tratar_colunas_oficial(recorte_carteira_oficial)
            recorte_carteira_glorian = tratar_colunas_glorian(recorte_carteira_glorian)

            g_codigos = set(recorte_carteira_glorian['Código'].unique())
            r_codigos = set(recorte_carteira_oficial['Código'].unique())

            t_codigos =  g_codigos | r_codigos
            

            for cod in g_codigos - r_codigos:
                t_codigos.remove(cod)
                msg = f"[!] Não foi encontrato o contrato de número {cod} na carteira oficial"
                df_mensagens.append({'Mensagem':msg}, ignore_index=True)
                df_mensagens = pd.concat([df_mensagens, pd.DataFrame([{'Mensagem':msg}])], ignore_index=True)
                
            for cod in r_codigos - g_codigos:
                t_codigos.remove(cod)
                msg = f"[!] Não foi encontrato o contrato de número {cod} na carteira do Glorian"
                df_mensagens = pd.concat([df_mensagens, pd.DataFrame([{'Mensagem':msg}])], ignore_index=True)
            
            diferencas = {}
            for cod in t_codigos:
                operacao_oficial = recorte_carteira_oficial[recorte_carteira_oficial['Código'] == cod]
                operacao_glorian = recorte_carteira_glorian[recorte_carteira_glorian['Código'] == cod]

                num_linhas_oficial = operacao_oficial.shape[0]
                num_linhas_glorian = operacao_glorian.shape[0]
                if num_linhas_oficial != num_linhas_glorian:
                    msg = f'[!] Diferenca no numero de linhas na operação {cod} ({num_linhas_oficial} na carteira e {num_linhas_glorian} no glorian)'
                    df_mensagens = pd.concat([df_mensagens, pd.DataFrame([{'Mensagem':msg}])], ignore_index=True)
                    continue

                for direcao in operacao_oficial['Operação'].unique():
                    cond = operacao_oficial['Operação'] == direcao
                    operacao_oficial2 = operacao_oficial[cond]
                    
                    cond = operacao_glorian['Operação'] == direcao
                    operacao_glorian2 = operacao_glorian[cond]
                    
                    if operacao_glorian2.shape[0] != operacao_oficial2.shape[0]:
                        msg = f'[!] Diferenca no numero de linhas na operação {cod}-{direcao} ({operacao_oficial2.shape[0]} na carteira e {operacao_glorian2.shape[0]} no glorian)'
                        df_mensagens = pd.concat([df_mensagens, pd.DataFrame([{'Mensagem':msg}])], ignore_index=True)
                        continue

                    operacao_oficial2 = operacao_oficial2.sort_values(['Mês','MWh'])
                    operacao_glorian2 = operacao_glorian2.sort_values(['inicio_fornecimento','Qtd vigente (MWh)'])

                    for i in range(operacao_oficial2.shape[0]):
                        diff = comparar_linhas(operacao_glorian2.iloc[i], operacao_oficial2.iloc[i])

                        if len(diff):
                            codigo = str(operacao_oficial2.iloc[i]["Código"])
                            diferencas[f'{codigo}'] = diff
                            df_diferencas.loc[f"{codigo}_{operacao_oficial2.iloc[i].name+2}_r", 'Código | Código'] = codigo
                            df_diferencas.loc[f"{codigo}_{operacao_glorian2.iloc[i].name+2}_g", 'Código | Código'] = codigo
            
            saida_diferencas()
            rotulo_status.configure(text="Concluído  ",image=icone_verde, compound="right")
            
        else:
            janela.update_idletasks()
            rotulo_status.configure(text="Data selecionada, não está presente no relatório!  ",image=icone_vermelho, compound="right")
            
            
def compara_carteira_glorian_peloMes_Ano(df_carteira_oficial,df_carteira_glorian, mesSelecionado, anoSelecionado):
    
        janela.update_idletasks()
        
        global df_diferencas
        global df_mensagens
        
        df_carteira_oficial['Ano'] = df_carteira_oficial['Mês'].dt.year
        df_carteira_oficial['MêsAtualizado'] = df_carteira_oficial['Mês'].dt.month
        
        
        filtro1 = df_carteira_oficial['MêsAtualizado']==mesSelecionado
        filtro2 = df_carteira_oficial['Ano']==anoSelecionado 
        
        filtro3 = df_carteira_glorian['Mês']==mesSelecionado
        filtro4 = df_carteira_glorian['Ano']==anoSelecionado 
                    
        recorte_carteira_oficial = df_carteira_oficial[filtro1 & filtro2].copy()
        recorte_carteira_glorian = df_carteira_glorian[filtro3 & filtro4].copy()
        
        if not recorte_carteira_glorian.empty and not recorte_carteira_oficial.empty:

            recorte_carteira_oficial= tratar_colunas_oficial(recorte_carteira_oficial)
            recorte_carteira_glorian = tratar_colunas_glorian(recorte_carteira_glorian)

            g_codigos = set(recorte_carteira_glorian['Código'].unique())
            r_codigos = set(recorte_carteira_oficial['Código'].unique())
            
            
            t_codigos =  g_codigos | r_codigos

            for cod in g_codigos - r_codigos:
                t_codigos.remove(cod)
                msg = f"[!] Não foi encontrato o contrato de número {cod} na carteira oficial"
                df_mensagens.append({'Mensagem':msg}, ignore_index=True)
                df_mensagens = pd.concat([df_mensagens, pd.DataFrame([{'Mensagem':msg}])], ignore_index=True)
                
            for cod in r_codigos - g_codigos:
                t_codigos.remove(cod)
                msg = f"[!] Não foi encontrato o contrato de número {cod} na carteira do Glorian"
                df_mensagens = pd.concat([df_mensagens, pd.DataFrame([{'Mensagem':msg}])], ignore_index=True)
            
            diferencas = {}
            for cod in t_codigos:
                operacao_oficial = recorte_carteira_oficial[recorte_carteira_oficial['Código'] == cod]
                operacao_glorian = recorte_carteira_glorian[recorte_carteira_glorian['Código'] == cod]

                num_linhas_oficial = operacao_oficial.shape[0]
                num_linhas_glorian = operacao_glorian.shape[0]
                if num_linhas_oficial != num_linhas_glorian:
                    msg = f'[!] Diferenca no numero de linhas na operação {cod} ({num_linhas_oficial} na carteira e {num_linhas_glorian} no glorian)'
                    df_mensagens = pd.concat([df_mensagens, pd.DataFrame([{'Mensagem':msg}])], ignore_index=True)
                    continue

                for direcao in operacao_oficial['Operação'].unique():
                    cond = operacao_oficial['Operação'] == direcao
                    operacao_oficial2 = operacao_oficial[cond]
                    
                    cond = operacao_glorian['Operação'] == direcao
                    operacao_glorian2 = operacao_glorian[cond]
                    
                    if operacao_glorian2.shape[0] != operacao_oficial2.shape[0]:
                        msg = f'[!] Diferenca no numero de linhas na operação {cod}-{direcao} ({operacao_oficial2.shape[0]} na carteira e {operacao_glorian2.shape[0]} no glorian)'
                        df_mensagens = pd.concat([df_mensagens, pd.DataFrame([{'Mensagem':msg}])], ignore_index=True)
                        continue

                    operacao_oficial2 = operacao_oficial2.sort_values(['Mês','MWh'])
                    operacao_glorian2 = operacao_glorian2.sort_values(['inicio_fornecimento','Qtd vigente (MWh)'])

                    for i in range(operacao_oficial2.shape[0]):
                        diff = comparar_linhas(operacao_glorian2.iloc[i], operacao_oficial2.iloc[i])

                        if len(diff):
                            codigo = str(operacao_oficial2.iloc[i]["Código"])
                            diferencas[f'{codigo}'] = diff
                            df_diferencas.loc[f"{codigo}_{operacao_oficial2.iloc[i].name+2}_r", 'Código | Código'] = codigo
                            df_diferencas.loc[f"{codigo}_{operacao_glorian2.iloc[i].name+2}_g", 'Código | Código'] = codigo
            
            saida_diferencas()
            rotulo_status.configure(text="Concluído  ",image=icone_verde, compound="right")
            
        else:
            janela.update_idletasks()
            rotulo_status.configure(text="Data selecionada, não está presente no relatório!  ",image=icone_vermelho, compound="right")
            
        
def compara_glorian_glorian(planilha_d0,planilha_d1):
        resetando_statusAplicacao()
        global path_carteira_glorian_3
        global path_carteira_glorian_4
    

        planilha_d0 = tratar_colunas_glorian(planilha_d0)
        planilha_d1 = tratar_colunas_glorian(planilha_d1)
    
        janela.update_idletasks()
        
        path_saida = os.path.abspath('saida')
                    
        if not os.path.exists(path_saida):
            os.makedirs(path_saida)

        file = os.path.basename(path_carteira_glorian_3)
        str_date = file.split("_")[2][0:8]
        str_date_d0 = pd.to_datetime(str_date, format="%Y%m%d").strftime("%d-%b-%y")

        file = os.path.basename(path_carteira_glorian_4)
        str_date = file.split("_")[2][0:8]
        str_date_d1 = pd.to_datetime(str_date, format="%Y%m%d").strftime("%d-%b-%y")

                    
        dif_net_mes_d0_d1=diferenca_Net_glorian(planilha_d0,planilha_d1,'mes',str_date_d0,str_date_d1)
        dif_net_ano_d0_d1=diferenca_Net_glorian(planilha_d0,planilha_d1,'Ano',str_date_d0,str_date_d1)

        path_dile_saida = os.path.join(path_saida, f'diferenca_planilhas_glorian.xlsx')
        
        try:
            writer = pd.ExcelWriter(f"{path_dile_saida}", engine='xlsxwriter')
        except:
            janela.update_idletasks()
            
            rotulo_status.configure(text="Data selecionada, não está presente no relatório!  ",image=icone_vermelho, compound="right")
            rotulo_status.configure(text="O arquivo diferenca_planilhas_glorian.xlsx está aberto, favor fechar.  ",image=icone_vermelho, compound="right")
            return
        
        writer = pd.ExcelWriter(f"{path_dile_saida}", engine='xlsxwriter')

        if not dif_net_mes_d0_d1.empty:
                    dif_net_mes_d0_d1=dif_net_mes_d0_d1.style.apply(lambda x: ['color: red' if v < 0 
                                                                            else 'color: green' for v in x])
                    

                    dif_net_ano_d0_d1=dif_net_ano_d0_d1.style.apply(lambda x: ['color: red' if v < 0 
                                                                            else 'color: green' for v in x])

                                                                            
                    dif_net_mes_d0_d1.to_excel(writer,sheet_name='NET com diferença',startcol=0)
                    dif_net_ano_d0_d1.to_excel(writer,sheet_name='NET com diferença',startcol=7)

        print(f"{len(dif_net_mes_d0_d1.index)} meses com diferença")
        operacoes_novas = diferenca_operacoes_glorian(planilha_d0,planilha_d1,'novas',str_date_d0,str_date_d1)
        if not operacoes_novas.empty:
            operacoes_novas.to_excel(writer,sheet_name='Operações Novas')

        operacoes_existentes = diferenca_operacoes_glorian(planilha_d0,planilha_d1,'existentes',str_date_d0,str_date_d1)
        if not operacoes_existentes.empty:
            operacoes_existentes.to_excel(writer,sheet_name='Operações Existentes')
            
        try:
            writer.save()
        except:
            writer.close()
            
        style_sheets(path_dile_saida)
        
        saida_aba2.delete(0, tk.END)  # Limpa o campo de entrada
        saida_aba2.insert(0, path_dile_saida)
        rotulo_status.configure(text="Concluido  ", image=icone_verde, compound="right")
        janela.update_idletasks()



            
def comparar_linhas(linha_glorian, linha_oficial):
    
        
        global df_diferencas
        diff = []

        for col_ofic, col_glor in relacao.items():

            valor_glorian = linha_glorian[col_glor]
            valor_oficial = linha_oficial[col_ofic]

            if col_ofic == 'CNPJ':
                if str(valor_oficial) == '0x2a':
                    valor_oficial = valor_glorian
                    #0x00: '#NULL!',  # Intersection of two cell ranges is empty
                    #0x07: '#DIV/0!', # Division by zero
                    #0x0F: '#VALUE!', # Wrong type of operand
                    #0x17: '#REF!',   # Illegal or deleted cell reference
                    #0x1D: '#NAME?',  # Wrong function or range name
                    #0x24: '#NUM!',   # Value range overflow
                    #0x2A: '#N/A',    # Argument or function not available
                else:
                    valor_glorian = np.int64(re.sub(r'[\.\/-]','',linha_glorian['CNPJ']))

                    
            elif col_ofic == 'Contraparte':
                valor_glorian = re.sub(linha_glorian['CNPJ']+' - ','',linha_glorian['Contraparte'])
            
            elif col_ofic == 'MWmed Final':
                valor_glorian = linha_glorian[col_glor] / linha_oficial['Horas']
                valor_oficial = round(valor_oficial,3)
                valor_glorian = round(valor_glorian,3)


            elif col_ofic == 'Data Base':

                if (valor_glorian != valor_oficial.strftime('%b/%Y')) and str(valor_glorian) != 'nan':
                    ano = valor_oficial.date().year
                    mes_anterior = valor_oficial.date().month -1
                    if mes_anterior == 0:
                        mes_anterior = 12
                        ano = ano - 1

                    valor_mes_anterior = datetime(ano,mes_anterior,1)

                    if valor_glorian == valor_mes_anterior.strftime('%b/%Y'):
                        valor_oficial = valor_mes_anterior 

                
                valor_oficial = valor_oficial.strftime('%b/%Y')
                    
                
            elif col_ofic == 'Perfil':
                if valor_glorian == 'CONV ESP' and valor_oficial == 'INC_0':
                    valor_oficial = valor_glorian
            
            # elif col_ofic == 'Meio de Fechamento':

            #     if valor_glorian == '' and valor_oficial == 'Balcao':
            #         valor_oficial = valor_glorian

            if valor_glorian != valor_oficial and str(valor_glorian) != str(valor_oficial):
                col = f'{col_ofic} | {col_glor}'
                diff.append([col, valor_oficial, valor_glorian])
                df_diferencas.loc[f"{linha_oficial['Código']}_{linha_oficial.name+2}_r", col] = valor_oficial
                df_diferencas.loc[f"{linha_glorian['Código']}_{linha_glorian.name+2}_g", col] = valor_glorian


        return diff

def colorir_celulas(path_dile_saida):
    workbook = load_workbook(filename=path_dile_saida)
    fill_color1 = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type="solid")

    for row_idx, row in enumerate(workbook['diferenças'].iter_rows(min_row=2), start=1):
        if row_idx % 4 in [1, 2]:
            for cell in row:
                cell.fill = fill_color1

    workbook.save(path_dile_saida)

def saida_diferencas():

        global selected_tab
            
        data_agora = datetime.now()
        path_saida = os.path.abspath('saida')
        
        if not os.path.exists(path_saida):
            os.makedirs(path_saida)
        
        if not df_diferencas['CNPJ | CNPJ'].isna().all():
            df_diferencas['CNPJ | CNPJ'] = df_diferencas['CNPJ | CNPJ'].fillna('').astype(str)
            df_diferencas['CNPJ | CNPJ'] = df_diferencas['CNPJ | CNPJ'].str.replace(r'(\d{1,2})(\d{3})(\d{3})(\d{4})(\d{2})', r'\1.\2.\3/\4-\5', regex=True)
        
        df_diferencas.dropna(axis=1, how='all', inplace=True)
        path_dile_saida = os.path.join(path_saida, f'diferenca_{data_agora.strftime("%Y%m%d%H%M%S")}.xlsx')
        writer = pd.ExcelWriter(path_dile_saida, engine='xlsxwriter')
        
        if df_mensagens.shape[0] > 0:
            df_mensagens.to_excel(writer, sheet_name='mensagens', index=False)
            for column in df_mensagens:
                column_length = max(df_mensagens[column].astype(str).map(len).max(), len(column))
                col_idx = df_mensagens.columns.get_loc(column)
                writer.sheets['mensagens'].set_column(col_idx, col_idx, column_length)
        df_diferencas.to_excel(writer, sheet_name='diferenças')
        
        for column in df_diferencas:
            column_length = max(df_diferencas[column].astype(str).map(len).max(), len(column))
            col_idx = df_diferencas.columns.get_loc(column)
            writer.sheets['diferenças'].set_column(col_idx, col_idx, column_length)
        try:
            writer.save()
        except:
            writer.close()

        colorir_celulas(path_dile_saida)

        print("Saida salva no arquivo :")
        print(path_dile_saida)  

        if selected_tab == 0:
            saida_aba1.delete(0, tk.END) 
            saida_aba1.insert(0, path_dile_saida)
            
        if selected_tab == 2:
            saida_aba3.delete(0, tk.END) 
            saida_aba3.insert(0, path_dile_saida)
        

def abrir_arquivo_excel(saida):
        arquivo_excel = saida.get() 
        subprocess.Popen(['start', 'excel', arquivo_excel], shell=True)
        
        
def criar_circulo_cor_fundo(cor, tamanho = (15, 15)):
    imagem = Image.new('RGBA', tamanho)
    desenho = ImageDraw.Draw(imagem)
    desenho.ellipse((0, 0, tamanho[0], tamanho[1]), fill=cor)
    return imagem


def resetando_statusAplicacao():
    rotulo_status.configure(text="",image="", compound="right")



def on_tab_selected(event):
    resetando_statusAplicacao()
    global selected_tab
    selected_tab = notebook.index(notebook.select())  # Obtém o índice da aba selecionada
    if selected_tab == 0:

        print("to aqui")
        
        botao_selecionar_arquivo1 = ctk.CTkButton(frame1, text="Selecionar Arquivo", command=lambda:selecionar_arquivo(entrada_caminho1,'oficial'))
        botao_selecionar_arquivo1.grid(row=0, column=2, padx=2, pady=2)
        botao_selecionar_arquivo2 = ctk.CTkButton(frame1, text="Selecionar Arquivo", command=lambda:selecionar_arquivo(entrada_caminho2))
        botao_selecionar_arquivo2.grid(row=1, column=2, padx=2, pady=2)
        botao_mostrar_dados1 = ctk.CTkButton(frame1, text="Comparar", command=mostrar_dados_carteira_glorian)
        botao_mostrar_dados1.grid(row=4, columnspan=3, padx=2, pady=2)
        botao_abrir_saida = ctk.CTkButton(frame1, text="Abrir Planilha", command=lambda:abrir_arquivo_excel(saida_aba1))
        botao_abrir_saida.grid(row=5, column=2, padx=2, pady=2)
        verificar_arquivos_entrada_aba_1e3()
    if selected_tab == 1:
        
        botao_selecionar_arquivo3 = ctk.CTkButton(frame2, text="Selecionar Arquivo", command=lambda:selecionar_arquivo(entrada_caminho3))
        botao_selecionar_arquivo3.grid(row=0, column=2, padx=2, pady=2)
        botao_selecionar_arquivo4 = ctk.CTkButton(frame2, text="Selecionar Arquivo", command=lambda:selecionar_arquivo(entrada_caminho4))
        botao_selecionar_arquivo4.grid(row=1, column=2, padx=2, pady=2)
        botao_mostrar_dados2 = ctk.CTkButton(frame2, text="Comparar", command=mostrar_dados_glorian_glorian)
        botao_mostrar_dados2.grid(row=3, columnspan=3, padx=2, pady=2)
        botao_abrir_saida = ctk.CTkButton(frame2, text="Abrir Planilha", command=lambda:abrir_arquivo_excel(saida_aba2))
        botao_abrir_saida.grid(row=4, column=2, padx=2, pady=2)
        verificar_arquivos_entrada_aba2()
        
    else:
        
        botao_selecionar_arquivo5 = ctk.CTkButton(frame3, text="Selecionar Arquivo", command=lambda:selecionar_arquivo(entrada_caminho5,'oficial'))
        botao_selecionar_arquivo5.grid(row=0, column=2, padx=2, pady=2)
        botao_selecionar_arquivo6 = ctk.CTkButton(frame3, text="Selecionar Arquivo", command=lambda:selecionar_arquivo(entrada_caminho6))
        botao_selecionar_arquivo6.grid(row=1, column=2, padx=2, pady=2)
        botao_mostrar_dados3 = ctk.CTkButton(frame3, text="Comparar", command=pegarValoresMesAno)
        botao_mostrar_dados3.grid(row=5, columnspan=3, padx=2, pady=2)
        botao_abrir_saida = ctk.CTkButton(frame3, text="Abrir Planilha", command=lambda:abrir_arquivo_excel(saida_aba3))
        botao_abrir_saida.grid(row=6, column=2, padx=2, pady=2)
        verificar_arquivos_entrada_aba_1e3()
        
    print(f"Aba selecionada: {selected_tab}")
    return selected_tab

def pegarValoresMesAno():
    mes_selecionado = month_var.get()
    ano_selecionado = year_var.get()
    data_formatada = f"{mes_selecionado}/{ano_selecionado}"
    mostrar_dados_carteira_glorian_pegandoValorPor_Mes_Ano(data_formatada)
    return data_formatada

if __name__ == "__main__":

    # Tamanho da aplicacao
        
        app_width=820
        app_height = 250 

        # 
        janela = tk.Tk()
        # janela.wm_iconbitmap("./faviIconraizenPower.ico")
        janela.title("Validador da carteira")
        janela.geometry(f"{app_width}x{app_height}")

        notebook = ttk.Notebook(janela)
        notebook.grid(row=0, column=0, sticky="nsew")
        notebook.grid_configure(columnspan=100)

        frame1 = ctk.CTkFrame(janela)
        frame2 = ctk.CTkFrame(janela)
        frame3 = ctk.CTkFrame(janela)
        
        # Adicione um evento para detectar a mudança de guia
        notebook.bind("<<NotebookTabChanged>>", on_tab_selected)
            
        rotulo_caminho1 = ctk.CTkLabel(frame1, text="Carteira [Raizen]:", font=("Helvetica", 14))
        rotulo_caminho1.grid(row=0, column=0, sticky="w")

        entrada_caminho1 = ctk.CTkEntry(frame1, width=app_width*0.63)
        entrada_caminho1.grid(row=0, column=1, padx=2, pady=2, sticky="ew")


        rotulo_caminho2 = ctk.CTkLabel(frame1, text="Portifólio_(V2) [Glorian]:", font=("Helvetica", 14))
        rotulo_caminho2.grid(row=1, column=0, sticky="w")

        entrada_caminho2 = ctk.CTkEntry(frame1, width=app_width*0.63)
        entrada_caminho2.grid(row=1, column=1, padx=2, pady=2, sticky="ew")

        rotulo_date1 = ctk.CTkLabel(frame1, text="Data Inicial:", font=("Helvetica", 14))
        rotulo_date1.grid(row=2, column=0)

        date1 = tkcalendar.DateEntry(frame1, date_pattern='dd/MM/yyyy')
        date1.grid(row=2, column=1, padx=2, pady=2, sticky="w")
        
        rotulo_date2 = ctk.CTkLabel(frame1, text="Data Final:", font=("Helvetica", 14))
        rotulo_date2.grid(row=3, column=0)

        date2 = tkcalendar.DateEntry(frame1, date_pattern='dd/MM/yyyy')
        date2.grid(row=3, column=1, padx=2, pady=2, sticky="w")

        # Crie círculos amarelos e verdes
        amarelo = criar_circulo_cor_fundo('yellow')
        verde = criar_circulo_cor_fundo('green')
        vermelho = criar_circulo_cor_fundo('red')

        # Crie objetos ImageTk para os círculos
        icone_amarelo = ImageTk.PhotoImage(amarelo)
        icone_verde = ImageTk.PhotoImage(verde)
        icone_vermelho = ImageTk.PhotoImage(vermelho)

        rotulo_saida = ctk.CTkLabel(frame1, text="Saída Diferença:", font=("Helvetica", 14))
        rotulo_saida.grid(row=5, column=0, sticky="w")

        saida_aba1 = ctk.CTkEntry(frame1)
        saida_aba1.grid(row=5, column=1, padx=2, pady=2, sticky="ew")
        
        
        rotulo_status = ctk.CTkLabel(janela, text="", font=("Helvetica", 14),text_color="black")
        rotulo_status.grid(row=6, column=97, padx=6, pady=1,sticky="se")
        
        rotulo_caminho3 = ctk.CTkLabel(frame2, text="Portifólio_(V2) [Glorian]:", font=("Helvetica", 14))
        rotulo_caminho3.grid(row=0, column=0, sticky="w")

        entrada_caminho3 = ctk.CTkEntry(frame2, width=app_width*0.63)
        entrada_caminho3.grid(row=0, column=1, padx=2, pady=2, sticky="ew")

        rotulo_caminho4 = ctk.CTkLabel(frame2, text="Portifólio_(V2) [Glorian]:", font=("Helvetica", 14))
        rotulo_caminho4.grid(row=1, column=0, sticky="w")

        entrada_caminho4 = ctk.CTkEntry(frame2, width=app_width*0.63)
        entrada_caminho4.grid(row=1, column=1, padx=2, pady=2, sticky="ew")

        rotulo_saida_aba2 = ctk.CTkLabel(frame2, text="Saída Diferença:", font=("Helvetica", 14))
        rotulo_saida_aba2.grid(row=4, column=0, sticky="w")

        saida_aba2 = ctk.CTkEntry(frame2)
        saida_aba2.grid(row=4, column=1, padx=2, pady=2, sticky="ew")
        
        
        rotulo_caminho5 = ctk.CTkLabel(frame3, text="Carteira [Raizen]:", font=("Helvetica", 14))
        rotulo_caminho5.grid(row=0, column=0, sticky="w")

        entrada_caminho5 = ctk.CTkEntry(frame3, width=app_width*0.63)
        entrada_caminho5.grid(row=0, column=1, padx=2, pady=2, sticky="ew")


        rotulo_caminho6 = ctk.CTkLabel(frame3, text="Portifólio_(V2) [Glorian]:", font=("Helvetica", 14))
        rotulo_caminho6.grid(row=1, column=0, sticky="w")

        entrada_caminho6 = ctk.CTkEntry(frame3, width=app_width*0.63)
        entrada_caminho6.grid(row=1, column=1, padx=2, pady=2, sticky="ew")
    
        month_label = ctk.CTkLabel(frame3, text="Mês:", font=("Helvetica", 14))
        month_label.grid(row=2, column=0)
        month_var = tk.StringVar()
        month_var.set("01")
        month_spinbox = tk.Spinbox(frame3, from_=1, to=12, textvariable=month_var, wrap=True, state="readonly")
        month_spinbox.grid(row=2, column=1,sticky='w')

        year_label = ctk.CTkLabel(frame3, text="Ano:", font=("Helvetica", 14))
        year_label.grid(row=4, column=0)

        year_var = tk.StringVar()
        year_var.set("2023")
        year_spinbox = tk.Spinbox(frame3, from_=2000, to=2100, textvariable=year_var, wrap=True, state="readonly")
        year_spinbox.grid(row=4, column=1,sticky='w')        
        
        rotulo_saida_aba3 = ctk.CTkLabel(frame3, text="Saída Diferença:", font=("Helvetica", 14))
        rotulo_saida_aba3.grid(row=6, column=0, sticky="w")

        saida_aba3 = ctk.CTkEntry(frame3)
        saida_aba3.grid(row=6, column=1, padx=2, pady=2, sticky="ew")
        
        notebook.add(frame1, text='Carteira - Glorian')
        notebook.add(frame2, text='Glorian - Glorian')
        notebook.add(frame3, text='Ano/Mês Carteira - Glorian')

        # Inicia a interface gráfica
        janela.mainloop()
    

